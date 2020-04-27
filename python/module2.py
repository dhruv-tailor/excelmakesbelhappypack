import parradoxReader as PR
import pathlib 
import os, re
import json
import random
from openpyxl import Workbook 
from IdeaGroup import *
import string
import localisation as singeltonloc
localisation = singeltonloc.localisation.getInstance()
class Policies:
	def __init__(self,name = "Empty idea/error" , type="", modifier=[] , allow = {}, potential= {}, json = {}):
		self.name = name
		self.englishname = localisation[name]
		self.type = type
		self.allow1 = ""
		self.allow2 = ""
		self.allow = allow
		self.potential = potential
		self.modifier = modifier
		self.json = json
		self.parseAllows()
		return
	def __repr__(self):
		ret = "Policies="+self.name + ": Type=" + self.type + " " + "modifiers:" + dict(self.modifier).__str__()
		return ret
	def parseAllows(self):
		allow = self.allow.copy()
		if 'calc_true_if' in allow : del allow["calc_true_if"]
		if 'NOT' in allow:
			del allow['NOT']
		if 'hidden_trigger' in allow: del allow["hidden_trigger"]
		if 'current_age' in allow : del allow['current_age']
		if len(allow) > 2:
			self.allow1 = "Error"
			self.allow2 = "Error more than 2 cond"
		elif len(allow) == 2:
			itera = iter(allow) 
			self.allow1 = self.parseAllowsHelper(allow[next(itera)])
			self.allow2 = self.parseAllowsHelper(allow[next(itera)])
		elif len(allow) == 1:
			if 'full_idea_group' in allow:
				if len(allow['full_idea_group']) == 2: 
					self.allow1 = self.parseAllowsHelper(allow['full_idea_group'][0])
					self.allow2 = self.parseAllowsHelper(allow['full_idea_group'][1])
					return
			try:
				itera = iter(allow) 
				ne = next(iter(allow) )
				self.allow1 = self.parseAllowsHelper(allow[ne])
				#self.allow2 = self.parseAllowsHelper(allow[ne])
			except KeyError:
				self.allow1 = str(allow[ne])
		else:
			pass
	def parseAllowsHelper(self, item):
		if type(item) == str:
			return self.findEngelishName(item)
		if type(item) == dict:
			if 'full_idea_group' in item:
				return self.findEngelishName(str(item['full_idea_group']))
			if 'OR' in item:
				return "OR group of legth {}".format(len(item['OR']['full_idea_group']))
			return str(item)
		if  type(item) == list:
			try:
				return self.findEngelishName(", ".join(item))
			except:
				return self.findEngelishName(str(item))
		return item
	def findEngelishName(self,item):
		regex = "([\w_0]+)"
		for match in re.findall(regex, item):
			if localisation[match] is not None:
				item = item.replace(match,localisation[match])
		return item

				
		
class PoliciesFactory:
	def ParseToJson(file):
		""" parse a file and return json"""
		if not pathlib.Path(file).exists():
				print('ERROR: Unable to find file: ' + str(file))
				raise FileNotFoundError 
		if pathlib.Path(file).suffix != ".txt":
			print('ERROR: File not a .txt: ' + str(file))
		return PR.decode(pathlib.Path(file).as_posix(),False,True)
	def separateKeys(json, keys = ["monarch_power" , "potential", "ai_will_do", "allow"]):
		""" Sperate the special keys"""
		keysDict = {}
		if type(keys) == str:
			if keys in json: del my_dict['key']
		if type(keys) == list:
			for ideaName in json:
				keysDict[ideaName] = {}
				for item in keys:
					if item in json[ideaName]: keysDict[ideaName][item] = json[ideaName].pop(item)
		return json , keysDict
	def makePolicies(json):
		ret = {}
		group, specKeys = PoliciesFactory.separateKeys(json)
		for name in group:
			try:
				modifier = group[name]
				ret[name] = Policies(name, type=specKeys[name]["monarch_power"],
							modifier=list(modifier.keys()),
							allow= specKeys[name].pop('allow', None),
							potential=specKeys[name].pop('potential', None),
							json=json[name] )
			except KeyError as err:
				print(err.args, name)
				raise err
		return ret
	def HandleDuplicats(json):
		names = []
		duplicates = []
		for name in json:
			if type(json[name]) == list:
				names.append(name)
		for name in names:
			letterItter = iter(list(string.ascii_lowercase))
			for duplicate in json.pop(name):
				json[name + next(letterItter)] = duplicate
		return json
	def ParseFromFile(file):
		out = PoliciesFactory.ParseToJson(file)
		out = PoliciesFactory.HandleDuplicats(out)
		return PoliciesFactory.makePolicies(out)

import psutil

def has_handle(fpath):
    for proc in psutil.process_iter():
        try:
            for item in proc.open_files():
                if fpath == item.path:
                    return True
        except Exception:
            pass

    return False

def test1():
	ModFolder = pathlib.PureWindowsPath(r"E:\Melle\Documents\Paradox Interactive\Europa Universalis IV\mod\GitBranch")
	files_ideas = [r"common\ideas\00_admin_ideas.txt" , r"common\ideas\00_dip_ideas.txt", r"common\ideas\00_mil_ideas.txt" ]
	files = [f for f in pathlib.Path(ModFolder /"common\\policies").iterdir() if pathlib.Path(f).suffix == ".txt"]
	out = []
	for f in files:
		f = ModFolder / pathlib.Path(f)
		if pathlib.Path(f).suffix == ".txt":
			data = PoliciesFactory.ParseFromFile(f)
			out.append(data)
			
	
	data = out[0]
	for f in out[1:]:
		data.update(f) #merge all the idea in dicts toghter	
	data
	print("making new workbook")
	workbook = Workbook()
	sheet = workbook.active
	x = 1
	y = 0
	titles = [k for k, val in Policies().__dict__.items() if not str(hex(id(val))) in str(val)]
	lengthsCollums = []
	for idx, title in enumerate(titles):
		sheet.cell(1,idx+1).value = title
		lengthsCollums.append(len(title))
	for row, name in enumerate(data):
		for y, title in enumerate(titles):
			val =  str(getattr(data[name], title))
			sheet.cell(row=row+2, column=y+1).value = val
			lengthsCollums[y] = max(len(val),lengthsCollums[y])
	from openpyxl.utils import get_column_letter
	for i, column_width in enumerate(lengthsCollums):
		sheet.column_dimensions[get_column_letter(i+1)].width = min(max(6 ,column_width), 60)


	print("saving to disk")
	fileName = "Policies.xlsx"
	fileName = pathlib.Path(fileName).resolve()
	repeat = True
	while(repeat):
		try:
			workbook.save(filename=fileName)
			repeat = False
		except:
			print("excel is open please close Policies.xlsx")
			os.system('pause')
	print(fileName)
	
	
if __name__ == "__main__":
	test1()


